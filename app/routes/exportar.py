from flask import Blueprint, request, jsonify, send_file
from ..models import db, Cliente, Barbero, Servicio, Cobro
from datetime import datetime
from io import BytesIO


def get_exportar_bp():
    bp = Blueprint('exportar', __name__)
    
    @bp.route('/generar', methods=['GET'])
    def generar_reporte():
        try:
            export_type = request.args.get('type', 'pdf')
            report_type = request.args.get('reportType', 'diario')
            fecha_inicio = request.args.get('fechaInicio', '')
            fecha_fin = request.args.get('fechaFin', '')
            
            cobros = Cobro.query.filter(
                db.func.date(Cobro.fecha) >= fecha_inicio,
                db.func.date(Cobro.fecha) <= fecha_fin
            ).all() if fecha_inicio and fecha_fin else []
            
            if export_type == 'excel':
                return generar_excel(cobros, report_type, fecha_inicio, fecha_fin)
            else:
                return generar_pdf(cobros, report_type, fecha_inicio, fecha_fin)
        
        except Exception as e:
            return jsonify({'error': f'Error: {str(e)}'}), 500

    return bp


def generar_excel(cobros, report_type, fecha_inicio, fecha_fin):
    try:
        from openpyxl.styles import Font, PatternFill
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte'
        
        headers = ['Fecha', 'Cliente', 'Barbero', 'Servicio', 'Monto']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
        
        total = 0
        for cobro in cobros:
            cliente = Cliente.query.get(cobro.cliente_id)
            barbero = Barbero.query.get(cobro.barbero_id)
            servicio = Servicio.query.get(cobro.servicio_id)
            
            ws.append([
                str(cobro.fecha.date()) if cobro.fecha else '',
                cliente.nombre if cliente else 'N/A',
                barbero.nombre if barbero else 'N/A',
                servicio.nombre if servicio else 'N/A',
                float(cobro.monto)
            ])
            total += float(cobro.monto)
        
        ws.append(['', '', '', 'TOTAL:', total])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reporte_{report_type}_{fecha_inicio}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'error': f'Error Excel: {str(e)}'}), 500


def generar_pdf(cobros, report_type, fecha_inicio, fecha_fin):
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#FF6B35'),
            spaceAfter=12,
            alignment=1
        )
        elements.append(Paragraph('BARBERÍA JORDAN', title_style))
        elements.append(Paragraph(f'Reporte {report_type.upper()}', styles['Heading2']))
        elements.append(Spacer(1, 0.3*inch))
        
        info_text = f'Período: {fecha_inicio} al {fecha_fin}<br/>Generado: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
        elements.append(Paragraph(info_text, styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        data = [['Fecha', 'Cliente', 'Barbero', 'Servicio', 'Monto']]
        total = 0
        
        for cobro in cobros:
            cliente = Cliente.query.get(cobro.cliente_id)
            barbero = Barbero.query.get(cobro.barbero_id)
            servicio = Servicio.query.get(cobro.servicio_id)
            
            monto_str = f"${float(cobro.monto):.2f}"
            data.append([
                str(cobro.fecha.date()) if cobro.fecha else '',
                cliente.nombre if cliente else 'N/A',
                barbero.nombre if barbero else 'N/A',
                servicio.nombre if servicio else 'N/A',
                monto_str
            ])
            total += float(cobro.monto)
        
        total_str = f"${total:.2f}"
        data.append(['', '', '', 'TOTAL', total_str])
        
        table = Table(data, colWidths=[1.2*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B35')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'reporte_{report_type}_{fecha_inicio}.pdf'
        )
    
    except Exception as e:
        return jsonify({'error': f'Error PDF: {str(e)}'}), 500
